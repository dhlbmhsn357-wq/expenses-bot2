import telebot
import openpyxl
import os
import datetime

TOKEN = "8792180324:AAGMO_GAgdKFlKSN9hI_u5OFYKQKTlDIZsg"
CHAT_ID = 5169632234
EXCEL_FILE = "مصاريفي.xlsx"

bot = telebot.TeleBot(TOKEN)

def get_workbook():
    if os.path.exists(EXCEL_FILE):
        return openpyxl.load_workbook(EXCEL_FILE)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "المصاريف"
    ws.append(["التاريخ", "الوقت", "النوع", "المبلغ", "ملاحظة"])
    wb.save(EXCEL_FILE)
    return wb

def save_expense(category, amount, note=""):
    wb = get_workbook()
    ws = wb.active
    now = datetime.datetime.now()
    ws.append([
        now.strftime("%Y-%m-%d"),
        now.strftime("%H:%M"),
        category,
        amount,
        note
    ])
    wb.save(EXCEL_FILE)

def get_weekly_report():
    wb = get_workbook()
    ws = wb.active
    today = datetime.date.today()
    week_ago = today - datetime.timedelta(days=7)
    total = 0
    categories = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            try:
                date = datetime.datetime.strptime(str(row[0]), "%Y-%m-%d").date()
                if week_ago <= date <= today:
                    amount = float(row[3])
                    total += amount
                    cat = row[2]
                    categories[cat] = categories.get(cat, 0) + amount
            except:
                pass
    return total, categories

def get_monthly_report():
    wb = get_workbook()
    ws = wb.active
    today = datetime.date.today()
    total = 0
    categories = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            try:
                date = datetime.datetime.strptime(str(row[0]), "%Y-%m-%d").date()
                if date.month == today.month and date.year == today.year:
                    amount = float(row[3])
                    total += amount
                    cat = row[2]
                    categories[cat] = categories.get(cat, 0) + amount
            except:
                pass
    return total, categories

def delete_by_category(category):
    wb = get_workbook()
    ws = wb.active
    rows_to_delete = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[2] and row[2] == category:
            rows_to_delete.append(i)
    for row_num in reversed(rows_to_delete):
        ws.delete_rows(row_num)
    wb.save(EXCEL_FILE)
    return len(rows_to_delete)

@bot.message_handler(commands=['start'])
def start(message):
    if message.chat.id != CHAT_ID:
        return
    bot.reply_to(message, """
مرحباً! 💰 أنا بوت متابعة مصاريفك

لتسجيل مصروف اكتب:
قهوة 25

الأوامر:
/اسبوع
/شهر
/حذف اسم المصروف
/ملف
    """)

@bot.message_handler(commands=['اسبوع'])
def weekly(message):
    if message.chat.id != CHAT_ID:
        return
    total, cats = get_weekly_report()
    if total == 0:
        bot.reply_to(message, "لا توجد مصاريف هذا الأسبوع")
        return
    text = "تقرير الأسبوع:\n\n"
    for cat, amount in sorted(cats.items(), key=lambda x: x[1], reverse=True):
        text += f"• {cat}: {amount:.0f} ج\n"
    text += f"\nالإجمالي: {total:.0f} ج"
    bot.reply_to(message, text)

@bot.message_handler(commands=['شهر'])
def monthly(message):
    if message.chat.id != CHAT_ID:
        return
    total, cats = get_monthly_report()
    if total == 0:
        bot.reply_to(message, "لا توجد مصاريف هذا الشهر")
        return
    text = "تقرير الشهر:\n\n"
    for cat, amount in sorted(cats.items(), key=lambda x: x[1], reverse=True):
        text += f"• {cat}: {amount:.0f} ج\n"
    text += f"\nالإجمالي: {total:.0f} ج"
    bot.reply_to(message, text)

@bot.message_handler(commands=['حذف'])
def delete_cmd(message):
    if message.chat.id != CHAT_ID:
        return
    parts = message.text.split(maxsplit=1)
    if len(parts) < 2:
        bot.reply_to(message, "اكتب اسم المصروف بعد الأمر\nمثال: /حذف قهوة")
        return
    category = parts[1].strip()
    count = delete_by_category(category)
    if count > 0:
        bot.reply_to(message, f"تم حذف {count} سجل من نوع {category}")
    else:
        bot.reply_to(message, f"مش لاقي مصروف باسم {category}")

@bot.message_handler(commands=['ملف'])
def send_file(message):
    if message.chat.id != CHAT_ID:
        return
    if os.path.exists(EXCEL_FILE):
        with open(EXCEL_FILE, 'rb') as f:
            bot.send_document(message.chat.id, f, caption="ملف مصاريفك")
    else:
        bot.reply_to(message, "لا يوجد ملف بعد!")

@bot.message_handler(func=lambda m: True)
def handle_expense(message):
    if message.chat.id != CHAT_ID:
        return
    text = message.text.strip()
    parts = text.rsplit(' ', 1)
    if len(parts) == 2:
        try:
            category = parts[0].strip()
            amount = float(parts[1].strip())
            save_expense(category, amount)
            now = datetime.datetime.now().strftime("%H:%M")
            bot.reply_to(message, f"تم التسجيل!\n{category}\n{amount:.0f} ج\n{now}")
        except ValueError:
            bot.reply_to(message, "اكتب مثلاً: قهوة 25")
    else:
        bot.reply_to(message, "اكتب مثلاً: قهوة 25")

print("بوت المصاريف شغال!")
bot.polling(none_stop=True)
